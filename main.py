import openpyxl

inv_file = openpyxl.load_workbook("files/inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {} #empty dictionary
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row+1): #default [0,74) so it will throw error on default
    supplier_name = product_list.cell(product_row, 4).value # product_list.cell(row num, column num)
    inventory_num_of_items = int(product_list.cell(product_row, 2).value)
    inventory_cost_per_item = float(product_list.cell(product_row, 3).value)
    product_num = int(product_list.cell(product_row, 1).value)
    total_product_value = product_list.cell(product_row, 5)

    # print products per company
    if supplier_name in products_per_supplier:
        # products_per_supplier["key"] = "value"
        current_num_of_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_of_products+1
    else:
        products_per_supplier[supplier_name] = 1

    #calculation total value per supplier

    if supplier_name in total_value_per_supplier:
        # products_per_supplier["key"] = "value"
        current_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_value + inventory_num_of_items * inventory_cost_per_item
    else:
        total_value_per_supplier[supplier_name] = inventory_num_of_items * inventory_cost_per_item

    # products with inv under 10
    if inventory_num_of_items < 10:
        products_under_10_inv[product_num] = inventory_num_of_items

    # add total value per product
    total_product_value.value = inventory_num_of_items * inventory_cost_per_item

# save edited file but creating a new one instead of override
inv_file.save("files/edited_inventory.xlsx")

print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)
